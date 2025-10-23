"""Tests for trade export functionality."""

import pytest
from pathlib import Path
from datetime import datetime, timedelta
from decimal import Decimal
import csv
import tempfile
import shutil

from trade_tracker.models.trade import StockTrade, OptionTrade, TradeType, OptionType
from trade_tracker.analytics.pnl import PnLCalculator, PositionPnL
from trade_tracker.utils.export import TradeExporter


@pytest.fixture
def sample_trades():
    """Create sample trades for testing."""
    base_date = datetime(2024, 1, 1)

    return [
        StockTrade(
            symbol="AAPL",
            trade_type=TradeType.BUY,
            quantity=100,
            price=Decimal("150.00"),
            commission=Decimal("1.00"),
            trade_date=base_date,
            account_id=1
        ),
        StockTrade(
            symbol="AAPL",
            trade_type=TradeType.SELL,
            quantity=100,
            price=Decimal("165.00"),
            commission=Decimal("1.00"),
            trade_date=base_date + timedelta(days=10),
            account_id=1
        ),
        StockTrade(
            symbol="MSFT",
            trade_type=TradeType.BUY,
            quantity=50,
            price=Decimal("300.00"),
            commission=Decimal("1.00"),
            trade_date=base_date + timedelta(days=15),
            account_id=1
        ),
        StockTrade(
            symbol="MSFT",
            trade_type=TradeType.SELL,
            quantity=50,
            price=Decimal("320.00"),
            commission=Decimal("1.00"),
            trade_date=base_date + timedelta(days=30),
            account_id=1
        ),
    ]


@pytest.fixture
def sample_pnl_results():
    """Create sample P/L results for testing."""
    return [
        PositionPnL(
            symbol="AAPL",
            quantity=100,
            entry_price=Decimal("150.00"),
            exit_price=Decimal("165.00"),
            entry_date=datetime(2024, 1, 1),
            exit_date=datetime(2024, 1, 11),
            cost_basis=Decimal("15001.00"),
            proceeds=Decimal("16499.00"),
            realized_pnl=Decimal("1498.00"),
            return_percentage=Decimal("9.99"),
            holding_period_days=10
        ),
        PositionPnL(
            symbol="MSFT",
            quantity=50,
            entry_price=Decimal("300.00"),
            exit_price=Decimal("320.00"),
            entry_date=datetime(2024, 1, 16),
            exit_date=datetime(2024, 1, 31),
            cost_basis=Decimal("15001.00"),
            proceeds=Decimal("15999.00"),
            realized_pnl=Decimal("998.00"),
            return_percentage=Decimal("6.65"),
            holding_period_days=15
        ),
    ]


@pytest.fixture
def temp_export_dir():
    """Create temporary directory for export testing."""
    temp_dir = tempfile.mkdtemp()
    yield Path(temp_dir)
    shutil.rmtree(temp_dir)


class TestTradeExporter:
    """Test TradeExporter class."""

    def test_export_trades_to_csv(self, sample_trades, sample_pnl_results, temp_export_dir):
        """Test exporting trades to CSV."""
        exporter = TradeExporter()
        output_file = temp_export_dir / "trades.csv"

        exporter.export_trades_to_csv(sample_trades, sample_pnl_results, output_file)

        assert output_file.exists()

        # Verify CSV content
        with open(output_file, 'r') as f:
            reader = csv.DictReader(f)
            rows = list(reader)

            assert len(rows) == 2  # 2 trade pairs

            # Check first row
            assert rows[0]['symbol'] == 'AAPL'
            assert rows[0]['quantity'] == '100'
            assert rows[0]['entry_price'] == '150.00'
            assert rows[0]['exit_price'] == '165.00'
            assert rows[0]['realized_pnl'] == '1498.00'
            assert rows[0]['return_pct'] == '9.99'

            # Check second row
            assert rows[1]['symbol'] == 'MSFT'
            assert rows[1]['quantity'] == '50'

    def test_export_empty_trades_to_csv(self, temp_export_dir):
        """Test exporting empty trade list."""
        exporter = TradeExporter()
        output_file = temp_export_dir / "empty_trades.csv"

        exporter.export_trades_to_csv([], [], output_file)

        assert output_file.exists()

        # Should have header but no data rows
        with open(output_file, 'r') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            assert len(rows) == 0

    def test_export_trades_to_excel(self, sample_trades, sample_pnl_results, temp_export_dir):
        """Test exporting trades to Excel."""
        pytest.importorskip("openpyxl")

        exporter = TradeExporter()
        output_file = temp_export_dir / "trades.xlsx"

        exporter.export_trades_to_excel(sample_trades, sample_pnl_results, output_file)

        assert output_file.exists()

        # Verify Excel content
        import openpyxl
        wb = openpyxl.load_workbook(output_file)

        # Check Trade History sheet
        assert "Trade History" in wb.sheetnames
        ws = wb["Trade History"]

        # Check headers
        assert ws['A1'].value == 'Symbol'
        assert ws['B1'].value == 'Entry Date'
        assert ws['C1'].value == 'Exit Date'

        # Check first data row
        assert ws['A2'].value == 'AAPL'
        assert ws['D2'].value == 100  # quantity
        assert ws['I2'].value == 1498.00  # realized P/L

    def test_export_monthly_summary_to_csv(self, sample_pnl_results, temp_export_dir):
        """Test exporting monthly P/L summary."""
        exporter = TradeExporter()
        output_file = temp_export_dir / "monthly_summary.csv"

        exporter.export_monthly_summary_to_csv(sample_pnl_results, output_file)

        assert output_file.exists()

        # Verify CSV content
        with open(output_file, 'r') as f:
            reader = csv.DictReader(f)
            rows = list(reader)

            assert len(rows) == 1  # All trades in January 2024
            assert rows[0]['month'] == '2024-01'
            assert rows[0]['total_pnl'] == '2496.00'
            assert rows[0]['num_trades'] == '2'

    def test_export_tax_report(self, sample_trades, sample_pnl_results, temp_export_dir):
        """Test generating tax report."""
        exporter = TradeExporter()
        output_file = temp_export_dir / "tax_report.csv"

        exporter.export_tax_report(sample_trades, sample_pnl_results, 2024, output_file)

        assert output_file.exists()

        # Verify tax report content
        with open(output_file, 'r') as f:
            reader = csv.DictReader(f)
            rows = list(reader)

            assert len(rows) == 2

            # Check that it includes tax-relevant fields
            assert 'symbol' in rows[0]
            assert 'date_acquired' in rows[0]
            assert 'date_sold' in rows[0]
            assert 'cost_basis' in rows[0]
            assert 'proceeds' in rows[0]
            assert 'gain_loss' in rows[0]
            assert 'term' in rows[0]  # short-term or long-term

    def test_export_tax_report_filters_by_year(self, sample_trades, sample_pnl_results, temp_export_dir):
        """Test that tax report filters by year."""
        exporter = TradeExporter()
        output_file = temp_export_dir / "tax_report_2025.csv"

        # Export for 2025 (should be empty)
        exporter.export_tax_report(sample_trades, sample_pnl_results, 2025, output_file)

        assert output_file.exists()

        with open(output_file, 'r') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            assert len(rows) == 0  # No trades in 2025

    def test_export_with_option_trades(self, temp_export_dir):
        """Test exporting option trades."""
        option_buy = OptionTrade(
            symbol="TSLA",
            trade_type=TradeType.BUY_TO_OPEN,
            quantity=10,
            price=Decimal("5.00"),
            strike=Decimal("250.00"),
            expiry=datetime(2024, 3, 15).date(),
            option_type=OptionType.CALL,
            commission=Decimal("6.50"),
            trade_date=datetime(2024, 1, 1),
            account_id=1
        )

        option_sell = OptionTrade(
            symbol="TSLA",
            trade_type=TradeType.SELL_TO_CLOSE,
            quantity=10,
            price=Decimal("8.00"),
            strike=Decimal("250.00"),
            expiry=datetime(2024, 3, 15).date(),
            option_type=OptionType.CALL,
            commission=Decimal("6.50"),
            trade_date=datetime(2024, 1, 20),
            account_id=1
        )

        pnl = PositionPnL(
            symbol="TSLA",
            quantity=10,
            entry_price=Decimal("5.00"),
            exit_price=Decimal("8.00"),
            entry_date=datetime(2024, 1, 1),
            exit_date=datetime(2024, 1, 20),
            cost_basis=Decimal("506.50"),
            proceeds=Decimal("793.50"),
            realized_pnl=Decimal("287.00"),
            return_percentage=Decimal("56.66"),
            holding_period_days=19
        )

        exporter = TradeExporter()
        output_file = temp_export_dir / "option_trades.csv"

        exporter.export_trades_to_csv([option_buy, option_sell], [pnl], output_file)

        assert output_file.exists()

        with open(output_file, 'r') as f:
            reader = csv.DictReader(f)
            rows = list(reader)

            assert len(rows) == 1
            assert rows[0]['symbol'] == 'TSLA'
            assert 'option_type' in rows[0]
            assert 'strike' in rows[0]
