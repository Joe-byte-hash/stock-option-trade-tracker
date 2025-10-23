"""Export functionality for trade data."""

import csv
from pathlib import Path
from typing import List, Optional
from datetime import datetime
from decimal import Decimal

from trade_tracker.models.trade import Trade, OptionTrade
from trade_tracker.analytics.pnl import PositionPnL


class TradeExporter:
    """Export trade data to various formats."""

    def export_trades_to_csv(
        self,
        trades: List[Trade],
        pnl_results: List[PositionPnL],
        output_file: Path
    ) -> None:
        """
        Export trade history with P/L to CSV.

        Args:
            trades: List of trades
            pnl_results: List of P/L results
            output_file: Output CSV file path
        """
        fieldnames = [
            'symbol', 'entry_date', 'exit_date', 'quantity',
            'entry_price', 'exit_price', 'cost_basis', 'proceeds',
            'realized_pnl', 'return_pct', 'holding_days'
        ]

        # Add option-specific fields if any option trades present
        has_options = any(isinstance(t, OptionTrade) for t in trades)
        if has_options:
            fieldnames.extend(['option_type', 'strike', 'expiry'])

        with open(output_file, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for pnl in pnl_results:
                row = {
                    'symbol': pnl.symbol,
                    'entry_date': pnl.entry_date.strftime('%Y-%m-%d'),
                    'exit_date': pnl.exit_date.strftime('%Y-%m-%d'),
                    'quantity': pnl.quantity,
                    'entry_price': f'{pnl.entry_price:.2f}',
                    'exit_price': f'{pnl.exit_price:.2f}',
                    'cost_basis': f'{pnl.cost_basis:.2f}',
                    'proceeds': f'{pnl.proceeds:.2f}',
                    'realized_pnl': f'{pnl.realized_pnl:.2f}',
                    'return_pct': f'{pnl.return_percentage:.2f}',
                    'holding_days': pnl.holding_period_days
                }

                # Add option fields if applicable
                if has_options:
                    # Find matching trade to get option details
                    matching_trade = next(
                        (t for t in trades if isinstance(t, OptionTrade) and t.symbol == pnl.symbol),
                        None
                    )
                    if matching_trade:
                        row['option_type'] = matching_trade.option_type.value
                        row['strike'] = f'{matching_trade.strike:.2f}'
                        row['expiry'] = matching_trade.expiry.strftime('%Y-%m-%d')
                    else:
                        row['option_type'] = 'N/A'
                        row['strike'] = 'N/A'
                        row['expiry'] = 'N/A'

                writer.writerow(row)

    def export_trades_to_excel(
        self,
        trades: List[Trade],
        pnl_results: List[PositionPnL],
        output_file: Path
    ) -> None:
        """
        Export trade history with P/L to Excel.

        Args:
            trades: List of trades
            pnl_results: List of P/L results
            output_file: Output Excel file path
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trade History"

        # Headers
        headers = [
            'Symbol', 'Entry Date', 'Exit Date', 'Quantity',
            'Entry Price', 'Exit Price', 'Cost Basis', 'Proceeds',
            'Realized P/L', 'Return %', 'Holding Days'
        ]

        # Add option headers if needed
        has_options = any(isinstance(t, OptionTrade) for t in trades)
        if has_options:
            headers.extend(['Option Type', 'Strike', 'Expiry'])

        # Write headers with formatting
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_num, pnl in enumerate(pnl_results, start=2):
            ws.cell(row=row_num, column=1, value=pnl.symbol)
            ws.cell(row=row_num, column=2, value=pnl.entry_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=3, value=pnl.exit_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=4, value=pnl.quantity)
            ws.cell(row=row_num, column=5, value=float(pnl.entry_price))
            ws.cell(row=row_num, column=6, value=float(pnl.exit_price))
            ws.cell(row=row_num, column=7, value=float(pnl.cost_basis))
            ws.cell(row=row_num, column=8, value=float(pnl.proceeds))
            ws.cell(row=row_num, column=9, value=float(pnl.realized_pnl))
            ws.cell(row=row_num, column=10, value=float(pnl.return_percentage))
            ws.cell(row=row_num, column=11, value=pnl.holding_period_days)

            # Color P/L cell based on win/loss
            pnl_cell = ws.cell(row=row_num, column=9)
            if pnl.realized_pnl > 0:
                pnl_cell.font = Font(color="27AE60")  # Green
            elif pnl.realized_pnl < 0:
                pnl_cell.font = Font(color="E74C3C")  # Red

            # Add option data if applicable
            if has_options:
                matching_trade = next(
                    (t for t in trades if isinstance(t, OptionTrade) and t.symbol == pnl.symbol),
                    None
                )
                if matching_trade:
                    ws.cell(row=row_num, column=12, value=matching_trade.option_type.value)
                    ws.cell(row=row_num, column=13, value=float(matching_trade.strike))
                    ws.cell(row=row_num, column=14, value=matching_trade.expiry.strftime('%Y-%m-%d'))

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_file)

    def export_monthly_summary_to_csv(
        self,
        pnl_results: List[PositionPnL],
        output_file: Path
    ) -> None:
        """
        Export monthly P/L summary to CSV.

        Args:
            pnl_results: List of P/L results
            output_file: Output CSV file path
        """
        # Group by month
        monthly_data = {}
        for pnl in pnl_results:
            month_key = pnl.exit_date.strftime('%Y-%m')
            if month_key not in monthly_data:
                monthly_data[month_key] = {
                    'total_pnl': Decimal('0'),
                    'num_trades': 0,
                    'wins': 0,
                    'losses': 0
                }

            monthly_data[month_key]['total_pnl'] += pnl.realized_pnl
            monthly_data[month_key]['num_trades'] += 1

            if pnl.realized_pnl > 0:
                monthly_data[month_key]['wins'] += 1
            elif pnl.realized_pnl < 0:
                monthly_data[month_key]['losses'] += 1

        # Write to CSV
        with open(output_file, 'w', newline='') as f:
            fieldnames = ['month', 'total_pnl', 'num_trades', 'wins', 'losses', 'win_rate']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for month in sorted(monthly_data.keys()):
                data = monthly_data[month]
                win_rate = (data['wins'] / data['num_trades'] * 100) if data['num_trades'] > 0 else 0

                writer.writerow({
                    'month': month,
                    'total_pnl': f'{data["total_pnl"]:.2f}',
                    'num_trades': data['num_trades'],
                    'wins': data['wins'],
                    'losses': data['losses'],
                    'win_rate': f'{win_rate:.2f}'
                })

    def export_tax_report(
        self,
        trades: List[Trade],
        pnl_results: List[PositionPnL],
        tax_year: int,
        output_file: Path
    ) -> None:
        """
        Export tax report for a specific year.

        Args:
            trades: List of trades
            pnl_results: List of P/L results
            tax_year: Tax year to report
            output_file: Output CSV file path
        """
        # Filter trades by tax year (based on exit date)
        year_pnl_results = [
            pnl for pnl in pnl_results
            if pnl.exit_date.year == tax_year
        ]

        with open(output_file, 'w', newline='') as f:
            fieldnames = [
                'symbol', 'date_acquired', 'date_sold', 'quantity',
                'cost_basis', 'proceeds', 'gain_loss', 'term'
            ]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for pnl in year_pnl_results:
                # Determine short-term vs long-term (>365 days)
                term = 'Long-term' if pnl.holding_period_days > 365 else 'Short-term'

                writer.writerow({
                    'symbol': pnl.symbol,
                    'date_acquired': pnl.entry_date.strftime('%Y-%m-%d'),
                    'date_sold': pnl.exit_date.strftime('%Y-%m-%d'),
                    'quantity': pnl.quantity,
                    'cost_basis': f'{pnl.cost_basis:.2f}',
                    'proceeds': f'{pnl.proceeds:.2f}',
                    'gain_loss': f'{pnl.realized_pnl:.2f}',
                    'term': term
                })
